#!/usr/bin/env python3
"""
Scrape posts from r/F1Discussions for the "takes" classification dataset.

Follows the data-collection spec in PLANNING.md:
  Round 1: ~40 posts that contain a question mark (likely 'question'/discussion).
  Round 2: keep collecting posts as they appear, up to a total of ~240.

The collected samples are written to an .xlsx file so they are easy to look
over and clean by hand. Once cleaned, the sheet can be exported as a .csv.

Title and body are merged into a single piece of text (the "text" column);
they are never stored as separate fields.

Reddit blocks unauthenticated requests and now gates API-app creation behind a
policy review, so this pulls from the Arctic Shift archive instead -- a public,
credential-free Reddit data archive (the community successor to Pushshift). It
returns the same post objects as Reddit's own API.
"""

import sys
import time
import html

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

SUBREDDIT = "F1Discussions"
QUESTION_TARGET = 40      # Round 1: posts containing "?"
TOTAL_TARGET = 270        # Round 1 + Round 2 combined
PAGE_LIMIT = 100          # posts per request (Arctic Shift max)
REQUEST_PAUSE = 1.0       # seconds between requests (be polite)
OUTPUT_XLSX = "f1discussions_samples.xlsx"

API_URL = "https://arctic-shift.photon-reddit.com/api/posts/search"
USER_AGENT = "python:takemeter-f1-classifier:v1.0 (data collection for fine-tuning)"

REMOVED_MARKERS = {"[removed]", "[deleted]", ""}


def fetch_new_posts(max_pages=15):
    """Yield post dicts newest-first, paginating by the created_utc of the last post."""
    headers = {"User-Agent": USER_AGENT}
    before = None
    seen_ids = set()
    for page in range(max_pages):
        params = {"subreddit": SUBREDDIT, "limit": PAGE_LIMIT, "sort": "desc"}
        if before is not None:
            params["before"] = before
        try:
            resp = requests.get(API_URL, headers=headers, params=params, timeout=30)
        except requests.RequestException as e:
            print(f"  ! request error on page {page+1}: {e}", file=sys.stderr)
            break

        if resp.status_code != 200:
            print(f"  ! HTTP {resp.status_code} on page {page+1}; stopping.", file=sys.stderr)
            break

        posts = resp.json().get("data", [])
        if not posts:
            break

        for post in posts:
            pid = post.get("id")
            if pid and pid not in seen_ids:
                seen_ids.add(pid)
                yield post

        # paginate further back in time using the oldest post in this page
        before = posts[-1].get("created_utc")
        print(f"  fetched page {page+1} ({len(posts)} posts, {len(seen_ids)} unique so far)")
        if before is None or len(posts) < PAGE_LIMIT:
            break
        time.sleep(REQUEST_PAUSE)


def combine_text(post):
    """Merge title + body into one long string. Returns '' if nothing usable."""
    title = html.unescape((post.get("title") or "").strip())
    body = html.unescape((post.get("selftext") or "").strip())
    if body in REMOVED_MARKERS:
        body = ""
    if not title and not body:
        return ""
    if body:
        return f"{title}\n\n{body}"
    return title


def main():
    print(f"Scraping r/{SUBREDDIT} via Arctic Shift ...")

    questions = []   # round 1: contains "?"
    others = []      # round 2: everything else
    seen = set()

    for post in fetch_new_posts():
        text = combine_text(post)
        if not text:
            continue
        pid = post.get("id")
        if pid in seen:
            continue

        record = {
            "id": pid,
            "text": text,
            "score": post.get("score", 0),
            "num_comments": post.get("num_comments", 0),
            "url": "https://www.reddit.com" + post.get("permalink", ""),
            "created_utc": time.strftime(
                "%Y-%m-%d %H:%M", time.gmtime(post.get("created_utc", 0))
            ),
        }

        if "?" in text and len(questions) < QUESTION_TARGET:
            questions.append(record)
            seen.add(pid)
        elif len(questions) + len(others) < TOTAL_TARGET:
            others.append(record)
            seen.add(pid)

        if len(questions) + len(others) >= TOTAL_TARGET:
            break

    samples = questions + others
    print(f"\nCollected {len(samples)} samples "
          f"({len(questions)} with '?', {len(others)} others).")

    # ---- write xlsx ----
    wb = Workbook()
    ws = wb.active
    ws.title = "samples"

    headers = ["id", "label", "text", "round", "score", "num_comments", "created_utc", "url"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in questions:
        ws.append([r["id"], "", r["text"], "1 (question)", r["score"],
                   r["num_comments"], r["created_utc"], r["url"]])
    for r in others:
        ws.append([r["id"], "", r["text"], "2 (general)", r["score"],
                   r["num_comments"], r["created_utc"], r["url"]])

    # readable column widths + wrapped text for the body column
    widths = {"A": 10, "B": 16, "C": 90, "D": 14, "E": 8, "F": 13, "G": 18, "H": 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_XLSX)
    print(f"Saved -> {OUTPUT_XLSX}")
    print("\nNext steps:")
    print("  1. Open the file, fill in the 'label' column (question / argument),")
    print("     and delete noisy / off-topic ('unclear') rows.")
    print("  2. Save As -> CSV when you're done.")


if __name__ == "__main__":
    main()
